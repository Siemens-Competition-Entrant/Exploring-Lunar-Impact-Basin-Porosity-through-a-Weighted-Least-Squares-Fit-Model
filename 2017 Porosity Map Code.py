import numpy
import openpyxl
import matplotlib.pyplot as plt
import math
from openpyxl.cell import column_index_from_string
from openpyxl.cell import get_column_letter
from openpyxl.styles import PatternFill, Fill, Color, Style, Font
from openpyxl.cell import Cell

#region_porosity = numpy.array([input ("Input the array of porosity values in the region, separated by commas")])

'''openpyxl.workbook.save_workbook(region_name)
global_column_length = global_right_column - global_left_column + 1
number_of_pixels_across = global_column_length / patch_side_length
global_row_length = global_bottom_row - global_top_row + 1
number_of_pixels_down = global_row_length / patch_side_length'''

'''

porosity_wb = openpyxl.load_workbook("OpenpyxlPorosityMapTest.xlsx")
porosity_sheet = porosity_wb.get_sheet_by_name("PorosityMaps")

porosity_sheet.cell('A1').value = "Hotel ID"
porosity_sheet.cell('A1').font = Font(bold=True)

porosity_wb.save("OpenpyxlPorosityMapTest.xlsx")

print (porosity_sheet.cell('A1').value)'''

from openpyxl import Workbook

region_name = input ("Input the name of the region you wish to study and press the enter key")
number_of_pixels_across = int(input("Input the number of pixels across"))
number_of_pixels_down = int(input("Input the number of pixels down"))
region_porosity = numpy.array([input ("Input the array of porosity values in the region, separated by commas")])

wb = Workbook()
ws = wb.active
ws.title = "Test Openpyxl"

#ws.cell(row=3, column=1).value = "SomeValue2"
for row in range(1, number_of_pixels_down + 1):
    for column in range(1, number_of_pixels_across + 1):
        porosity_index = 1 + (row - 1) * number_of_pixels_across + column
        ws.cell(row = row, column = column).value = region_porosity[porosity_index]
wb.save(region_name + ".xlsx")


#porosity_sheet['A1'] = openpyxl.styles.PatternFill(start_color = '000050')
#porosity_sheet['A1'] = PatternFill(fill_type=None,
              #start_color='FFFFFFFF')

'''cell = porosity_sheet.cell('A1')
#cell.style.fill.fill_type = Fill.FILL_SOLID
cell.style.fill.start_color.index = Color.FFFF0000
redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000')
porosity_sheet['A1'].fill = redFill'''
'''porosity_sheet['A1'] = 1
porosity_sheet.cell(row=2, column=2).value = 2
porosity_sheet.cell(coordinate="C3").value = 3'''
rgb=[255,0,0]
color_string="".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])
porosity_sheet["A1"].fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
porosity_sheet['A1'].style = Style(fill=PatternFill(patternType='solid',
                                        fill_type='solid',
                                        fgColor=Color('C4C4C4')))
#porosity_sheet['A1'] = openpyxl.styles.PatternFill(bgColor="FFC7CE", fill_type="solid")
'''for i in range(1, int(number_of_pixels_down)):
    for j in range(1, int(number_of_pixels_across)):
        porosity_sheet.cell(row = i, column = j)
        pixel_number = number_of_pixels_across * i + j
        print(pixel_number)
        #if (pixel_number - 1) % patch_side_length > 0:
        if (pixel_number) <= (number_of_pixels_across * number_of_pixels_down):
            porosity_sheet_cell.value = region_porosity[pixel_number - 1]
        #if (pixel_number - 1) % patch_side_length == 0:
            #porosity_sheet_cell.value = region_porosity[pixel_number]
        if porosity_sheet_cell.value > 4:
            porosity_sheet [porosity_sheet_cell.value].fill = PatternFill(start_color='000000080',
                                  end_color='000000080',
                                  fill_type='solid')'''
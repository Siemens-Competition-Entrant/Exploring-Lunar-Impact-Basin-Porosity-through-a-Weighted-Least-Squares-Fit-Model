import numpy
import openpyxl
import matplotlib.pyplot as plt
import math
from openpyxl.cell import column_index_from_string
from openpyxl.cell import get_column_letter
from openpyxl.styles import PatternFill, Fill, Color, Style, Font
from openpyxl.cell import Cell
from openpyxl import Workbook

#Opens workbooks -- CHANGE THIS TO THE UPDATED DATA FILES!!!!!
'''grav_wb = openpyxl.load_workbook("GRAIL.xlsx", read_only=True)
topo_wb = openpyxl.load_workbook("LOLA.xlsx", read_only= True)
grain_density_wb = openpyxl.load_workbook("GrainDensity.xlsx", read_only= True)
grav_sheet = grav_wb.get_sheet_by_name("Gravity")
topo_sheet = topo_wb.get_sheet_by_name("Gravity-from-topography")
grain_density_sheet = grain_density_wb.get_sheet_by_name("GrainDensity")'''

#Takes user input of cells and information about the selected cells
region_name = input ("Input the name of the region you wish to study and press the enter key")
type_of_data = input("What kind of files: Basin, Rim, or Exterior?")
global_top_left = input("Input the coordinates of the top-left corner of the rectangular area you wish to study and press the enter key")
global_bottom_right = input("Input the coordinates of the bottom-right corner of rectangular area you wish to study and press the enter key")
'''number_of_weights = int(input("Input the number of weights you wish to use in the weighted least-squares fit (10 is suggested)"))
patch_side_length = int(input("Input the length of a patch side (20 is suggested, must be an even number)"))
number_of_km_per_pixel = float(input("Input the number of kilometers per Excel pixel"))'''
number_of_weights = 10
patch_side_length = 20
number_of_km_per_pixel = 5
'''grav_wb_name = input("Input the name of the gravity file you plan to use")
grav_sheet_name = input("Input the name of the gravity sheet you plan to use")
topo_wb_name = input("Input the name of the gravity-from-topography file you plan to use")
topo_sheet_name = input("Input the name of the gravity-from-topography sheet you plan to use")
grain_density_wb_name = input("Input the name of the grain density file you plan to use")
grain_density_sheet_name = input("Input the name of the grain density sheet you plan to use")
global_top_left = input("Input the coordinates of the top-left corner of the rectangular area you wish to study and press the enter key")
global_bottom_right = input("Input the coordinates of the bottom-right corner of rectangular area you wish to study and press the enter key")
number_of_weights = int(input("Input the number of weights you wish to use in the weighted least-squares fit (10 is suggested)"))
patch_side_length = int(input("Input the length of a patch side (20 is suggested, must be an even number)"))
number_of_pixels_across = int(input("Input the number of porosity pixels across (include overlaps)"))
number_of_pixels_down = int(input("Input the number of porosity pixels down (include overlaps)"))
number_of_km_per_pixel = float(input("Input the number of kilometers per Excel pixel"))'''



#Opens workbooks based on the name input
'''grav_wb = openpyxl.load_workbook(grav_wb_name + ".xlsx", read_only=True)
topo_wb = openpyxl.load_workbook(topo_wb_name + ".xlsx", read_only= True)
grain_density_wb = openpyxl.load_workbook(grain_density_wb_name + ".xlsx", read_only= True)
grav_sheet = grav_wb.get_sheet_by_name(grav_sheet_name)
topo_sheet = topo_wb.get_sheet_by_name(topo_sheet_name)
grain_density_sheet = grain_density_wb.get_sheet_by_name(grain_density_sheet_name)'''

grav_wb = openpyxl.load_workbook(region_name + "_" + type_of_data + "Gravity.xlsx", read_only=True)
topo_wb = openpyxl.load_workbook(region_name + "_" + type_of_data + "GravityFromTopography.xlsx", read_only=True)
grain_density_wb = openpyxl.load_workbook(region_name  + "_GrainDensity.xlsx", read_only=True)
grav_sheet = grav_wb.get_sheet_by_name("Sheet1")
topo_sheet = topo_wb.get_sheet_by_name("Sheet1")
grain_density_sheet = grain_density_wb.get_sheet_by_name("Sheet1")

#Creates starting columns
global_top_row = int(grav_sheet.cell(global_top_left).row)
global_left_column = int(grav_sheet.cell(global_top_left).column)
start_global_left_column = global_left_column
global_bottom_row = int(grav_sheet.cell(global_bottom_right).row)
global_right_column = int(grav_sheet.cell(global_bottom_right).column)
print("Calculating least-squares fits and region properties on the farside for the region you chose.  Please be patient as this may take up to several days.")

number_of_pixels_across = int((global_right_column - global_left_column + 1) / 10 - 1)
number_of_pixels_down = int((global_bottom_row - global_top_row + 1) / 10 - 1)

#Creates needed arrays for density and porosity
region_density = numpy.array([])
confidence_intervals = numpy.array([])
region_porosity_with_confidence_intervals = numpy.array([])
region_porosity = numpy.array([])

#Creates an array for datapoints, which will be modified to be weight numbers
datapoint_positions = numpy.array([])
datapoint_position_weights = numpy.array([])
for i in range (1, patch_side_length ** 2 + 1):
    datapoint_positions = numpy.append(datapoint_positions, i)

#Calculates values needed to determine weights based on the patch size and numbeer of weights needed
middle_of_patch_side = patch_side_length / 2
center_value_of_patch_side = middle_of_patch_side + .5
width_of_weight_rings = middle_of_patch_side / number_of_weights

#Creates an array of weight values based on the positions of the datapoints in the patch
for i in datapoint_positions:
    #Calculates the rows and columns a given datapoint is in
    if i % patch_side_length > 0:
        row = 1 + ((i / patch_side_length) - ((i % patch_side_length) / patch_side_length))
    if i % patch_side_length == 0:
        row = (i / patch_side_length)
    if i % patch_side_length > 0:
        column = i % patch_side_length
    if i % patch_side_length == 0:
        column = patch_side_length
    #Calculates how far the given datapoint is from the center of the patch
    row_from_center = abs(row - center_value_of_patch_side) + .5
    column_from_center = abs(column - center_value_of_patch_side) + .5
    #Creates square rings based on the distance of a datapoint from the patch (the ring closest to the center has the highest number)
    if row_from_center >= column_from_center:
        ring = 1 + middle_of_patch_side - row_from_center
    if row_from_center < column_from_center:
        ring = 1 + middle_of_patch_side - column_from_center
    #Figures out the number of rings needed based on the number of weights desired, and modifies the ring assigned to a datapoint based on the number of weights needed
    if ring % width_of_weight_rings > 0:
        ring_for_weight = 1 + (((ring - 1) / width_of_weight_rings) - (((ring - 1) % width_of_weight_rings) / width_of_weight_rings))
    if ring % width_of_weight_rings == 0:
        ring_for_weight = 1 + (((ring - width_of_weight_rings) / width_of_weight_rings))
    #Calculates a weight value based on the position of a datapoint and the number of rings needed and inputs it into an array
    position_weight = ring_for_weight / number_of_weights
    datapoint_position_weights = numpy.append(datapoint_position_weights, position_weight)

#plot_number = 1
#Calculates density and porosity in regions, spaced by 20 pixels
for global_top_row in range(global_top_row, global_bottom_row - (patch_side_length - 2), 10):
    for global_left_column in range(global_left_column, global_right_column - (patch_side_length - 2), 10):
#Preparing coordinates for analysis
        #To create coordinates for studied region
        left_column = get_column_letter(global_left_column)
        top_left_coordinates = str(str(left_column) + str(global_top_row))
        right_column = global_left_column + (patch_side_length - 1)
        right_column = get_column_letter(right_column)
        bottom_right_coordinates = str(str(right_column) + str(global_top_row + (patch_side_length - 1)))
        #print("Region from " + top_left_coordinates + " to " + bottom_right_coordinates + " -- plot number " + str(plot_number))

        #Retrieves data from datasets
        grav_data_input = tuple(grav_sheet[top_left_coordinates : bottom_right_coordinates])
        topo_data_input = tuple(topo_sheet[top_left_coordinates : bottom_right_coordinates])
        grain_density_data_input = tuple(grain_density_sheet[top_left_coordinates : bottom_right_coordinates])

#Calculating density based on weighted least-squares fit
        #Inputs x and y data
        y = numpy.array([])
        for rowOfCellObjects in grav_data_input:
            for cellObj in rowOfCellObjects:
                y_data = cellObj.value
                y = numpy.append(y, y_data)

        x = numpy.array([])
        for rowOfCellObjects in topo_data_input:
            for cellObj in rowOfCellObjects:
                x_data = cellObj.value
                x = numpy.append(x, x_data)

        #w_i = math.cos(median_latitude * math.pi / 180) * datapoint_position_weights
        #Renames weight array
        w_i = datapoint_position_weights

        #Weighted mean of x and y values
        sum_of_weighted_x_values = 0
        for i in range (0, len(x)):
            x_i = x[i]
            x_weighted_value = x_i * w_i[i]
            sum_of_weighted_x_values = sum_of_weighted_x_values + x_weighted_value

        sum_of_weighted_y_values = 0
        for i in range (0, len(y)):
            y_i = y[i]
            y_weighted_value = y_i * w_i[i]
            sum_of_weighted_y_values = sum_of_weighted_y_values + y_weighted_value

        weight_sum = numpy.sum([w_i])
        y_weighted_mean = sum_of_weighted_y_values / weight_sum
        x_weighted_mean = sum_of_weighted_x_values / weight_sum

        # Computes the sum of deviation in x and y values from the mean values, with the deviation in x multiplied by the deviation in y for a given datapoint
        sum_of_x_errors_times_y_errors = 0
        for i in range (0, len(x)):
            x_i = x[i]
            y_at_x = y[i]
            x_error = x_i - x_weighted_mean
            y_error = y_at_x - y_weighted_mean
            x_error_times_y_error = x_error * y_error
            weighted_x_error_times_y_error = x_error_times_y_error * w_i[i]
            sum_of_x_errors_times_y_errors = sum_of_x_errors_times_y_errors + weighted_x_error_times_y_error

        #Computes the sum of deviation in x values from the weighted mean and squares it
        sum_of_x_errors_squared = 0
        for i in range (0, len(x)):
            x_i = x[i]
            x_error = x_i - x_weighted_mean
            x_error_squared = x_error ** 2
            weighted_x_error_squared = x_error_squared * w_i[i]
            sum_of_x_errors_squared = sum_of_x_errors_squared + weighted_x_error_squared

        # Performs weighted least-squares fit
        density = sum_of_x_errors_times_y_errors / sum_of_x_errors_squared
        b = y_weighted_mean - density * x_weighted_mean
        region_density = numpy.append(region_density, density)

#Calculates standard error and confidence interval in density measurement
        # Computes the sum of squared differences in the actual y-value from the predicted y-value -- error in the line of best fit
        best_fit_error = 0
        for i in range (0, len(y)):
            difference_from_predicted = y[i] - (density * x[i] + b)
            difference_from_regression_squared = difference_from_predicted ** 2
            weighted_difference_from_predicted = w_i[i] * difference_from_regression_squared
            best_fit_error = best_fit_error + weighted_difference_from_predicted

        # Computes the area of the selected region on the Moon
        min_row = grav_sheet.cell(top_left_coordinates).row
        min_column = grav_sheet.cell(top_left_coordinates).column
        max_row = grav_sheet.cell(bottom_right_coordinates).row
        max_column = grav_sheet.cell(bottom_right_coordinates).column

        rows = max_row - min_row + 1
        columns = max_column - min_column + 1

        '''latitude_height = rows / 10
        longitude_width = columns / 10

        pi = math.pi
        #patch_area = latitude_height * longitude_width * ((1738 * pi / 180) ** 2) * math.cos(median_latitude * pi / 180)
        patch_area = latitude_height * longitude_width * ((1738 * pi / 180) ** 2)'''

        rows_in_km = rows* number_of_km_per_pixel
        columns_in_km = columns * number_of_km_per_pixel
        patch_area = rows_in_km * columns_in_km

        # Computes the standard error and creates the confidence interval
        standard_error = math.sqrt(best_fit_error / sum_of_x_errors_squared) / ((patch_area * 234726 / (3.8 * 10 ** 7))) #Check this
        # print("The standard error of the slope measurement is " + str(standard_error))
        standard_error = round(standard_error, 4)
        confidence_interval = standard_error * 2
        confidence_intervals = numpy.append(confidence_intervals, str(density) + " ± " + str(confidence_interval))
        #print("The density of the bedrock in this region is " + str(m) + " ± " + str(confidence_interval) + " g/cc")

#Converts density to porosity (with confidence interval)
        #Inputs average grain density in a pixel and averages it
        grain_density = numpy.array([])
        for rowOfCellObjects in grain_density_data_input:
            for cellObj in rowOfCellObjects:
                grain_density_data = cellObj.value / 1000
                grain_density = numpy.append(grain_density, grain_density_data)
        sum_of_weighted_grain_density_values = 0
        for i in range (0, len(grain_density)):
            grain_density_i = grain_density[i]
            sum_of_weighted_grain_density_values = sum_of_weighted_grain_density_values + (grain_density_i * w_i[i])

        grain_density_weighted_mean = sum_of_weighted_grain_density_values / weight_sum
        grain_density_weighted_mean = round(grain_density_weighted_mean, 3)

        #Calculates porosity and confidence interval based on density and grain density
        porosity = 1 - density / grain_density_weighted_mean
        percent_porosity = porosity * 100
        percent_porosity = round(percent_porosity, 3)
        region_porosity = numpy.append(region_porosity, percent_porosity)
        porosity_error = standard_error / grain_density_weighted_mean
        percent_porosity_error = porosity_error * 100
        percent_porosity_error = round(percent_porosity_error, 3)
        porosity_confidence_interval = percent_porosity_error * 2
        porosity_with_confidence_interval = str(percent_porosity) + " ± " + str(porosity_confidence_interval)
        region_porosity_with_confidence_intervals = numpy.append(region_porosity_with_confidence_intervals, porosity_with_confidence_interval)

#Shift studied region to next pixel
        '''start_longitude = round(start_longitude, 1)
        #Change this for each region

        if global_left_column >= global_right_column - (patch_side_length - 1):
            start_longitude = original_longitude

        if global_left_column < global_right_column - (patch_side_length - 1):
            start_longitude = start_longitude + 2'''

        if global_left_column >= global_right_column - (patch_side_length - 1):
            #median_latitude = median_latitude - 2
            global_left_column = start_global_left_column

        #plot_number = plot_number + 1

#Prints density and porosity arrays
print ("Density = " + str(confidence_intervals))
print ("Porosity = " + str(region_porosity))
print("Porosity with confidence interval = " + str(region_porosity_with_confidence_intervals))

#Creates workbook for porosity maps
porosity_wb = Workbook()
porosity_sheet = porosity_wb.active
porosity_sheet.title = region_name

#Creates fill colors.  Link to color converter: http://www.rapidtables.com/convert/color/rgb-to-hex.htm -- put FF before the 6-digit code the website gives you
blueFill_1 = PatternFill(start_color='FF000050',
                         end_color='FF000050',
                         fill_type='solid')
#blueFill_1 RGB code: 0,0,80
blueFill_2 = PatternFill (start_color='FF3333FF',
                         end_color='FF3333FF',
                         fill_type='solid')
#blueFill_2 RGB code: 51,51,255
blueFill_3 = PatternFill (start_color='FFA8A8FF',
                         end_color='FFA8A8FF',
                         fill_type='solid')
#blueFill_3 RGB code: 168,168,255
redFill_1 = PatternFill (start_color='FFFF4646',
                         end_color='FFFF4646',
                         fill_type='solid')
#redFill_1 RGB code: 255,70,70
redFill_2 = PatternFill(start_color='FFD40000',
                        end_color='FFD40000',
                        fill_type='solid')
#redFill_2 RGB code: 212,0,0
redFill_3 = PatternFill(start_color='FF8A0000',
                        end_color='FF8A0000',
                        fill_type='solid')
#redFill_3 RGB code: 138,0,0

#Creates porosity maps
for row in range(1, number_of_pixels_down + 1):
    for column in range(1, number_of_pixels_across + 1):
        #Prints porosity values in Excel workbook
        porosity_index = (row - 1) * number_of_pixels_across + column - 1
        porosity_sheet.cell(row = row, column = column).value = region_porosity[porosity_index]
        #Turns font white
        porosity_sheet.cell(row=row, column=column).font = Font(color='FFFFFFFF')

        #Colors cells based on porosity value
        if region_porosity[porosity_index] < 10:
            porosity_sheet.cell(row=row, column=column).fill = blueFill_1
        if 10 <= region_porosity[porosity_index] < 12:
            porosity_sheet.cell(row=row, column=column).fill = blueFill_2
        if 12 <= region_porosity[porosity_index] < 14:
            porosity_sheet.cell(row=row, column=column).fill = blueFill_3
        if 14 <= region_porosity[porosity_index] < 16:
            porosity_sheet.cell(row=row, column=column).fill = redFill_1
        if 16 <= region_porosity[porosity_index] < 18:
            porosity_sheet.cell(row=row, column=column).fill = redFill_2
        if region_porosity[porosity_index] >= 18:
            porosity_sheet.cell(row=row, column=column).fill = redFill_3

        #Rewrites cell values to include confidence intervals once the cells are colored
        porosity_sheet.cell(row = row, column = column).value = region_porosity_with_confidence_intervals[porosity_index]
#Saves the porosity map Excel file
porosity_wb.save(region_name + " (" + type_of_data + " Files).xlsx")


'''y = numpy.array([6, 8, 8, 10, 12])
x = numpy.array([1, 2, 3, 4, 5])
w_i = numpy.array([.1, .2, .3, .4, .5])
grain_density = numpy.array([1.1, 1.2, 1.3, 1.4, 1.5])
latitude_height = 2
longitude_width = 2
median_latitude = 45

for i in x:
    print (w_i[numpy.where(x == i)])

print (numpy.sum([w_i]))

sum_of_weighted_x_values = 0
for i in range (0, len(x)):
    x_i = x[i]
    sum_of_weighted_x_values = sum_of_weighted_x_values + x_i * w_i[i]

sum_of_weighted_y_values = 0
for i in range (0, len(y)):
    y_i = y[i]
    sum_of_weighted_y_values = sum_of_weighted_y_values + y_i * w_i[i]

weight_sum = numpy.sum([w_i])
y_weighted_mean = sum_of_weighted_y_values / weight_sum
x_weighted_mean = sum_of_weighted_x_values / weight_sum

# Computes the sum of squared deviation in x and y values from the mean values

sum_of_x_errors_times_y_errors = 0
for i in range (0, len(x)):
    x_i = x[i]
    y_at_x = y[i]
    x_error = x_i - x_weighted_mean
    y_error = y_at_x - y_weighted_mean
    x_error_times_y_error = x_error * y_error
    sum_of_x_errors_times_y_errors = sum_of_x_errors_times_y_errors + x_error_times_y_error * w_i[i]

sum_of_x_errors_squared = 0
for i in range (0, len(y)):
    y_i = y[i]
    y_error = y_i - y_weighted_mean
    y_error_squared = y_error ** 2
    sum_of_x_errors_squared = sum_of_x_errors_squared + y_error_squared * w_i[i]
# Performs weighted least-squares fit
density = sum_of_x_errors_times_y_errors / sum_of_x_errors_squared
print (density)
b = y_weighted_mean - density * x_weighted_mean

best_fit_error = 0
for i in range(0, len(y)):
    difference_from_predicted = y[i] - (density * x[i] + b)
    difference_from_regression_squared = difference_from_predicted ** 2
    weighted_difference_from_predicted = w_i[i] * difference_from_regression_squared
    best_fit_error = best_fit_error + weighted_difference_from_predicted
# Computes the area of the selected region on the Moon

pi = math.pi
patch_area = latitude_height * longitude_width * ((1738 * pi / 180) ** 2) * math.cos(median_latitude * pi / 180)

# Computes the standard error and creates the confidence interval
standard_error = math.sqrt(best_fit_error / sum_of_x_errors_squared) / ((patch_area * 241001 / (3.8 * 10 ** 7)))
# print("The standard error of the slope measurement is " + str(standard_error))
standard_error = round(standard_error, 4)
confidence_interval = standard_error * 2
print(confidence_interval)

grain_density_mean = numpy.mean(grain_density)
grain_density_mean = round(grain_density_mean, 3)
porosity = 1 - density / grain_density_mean
percent_porosity = porosity * 100
percent_porosity = round(percent_porosity, 3)
porosity_error = standard_error / grain_density_mean
porosity_error = round(porosity_error, 3)
porosity_confidence_interval = porosity_error * 2
print(str(percent_porosity) + " ± " + str(porosity_confidence_interval))'''



#ORIGINAL CODE -- DOES NOT HANDLE REPEAT VALUES IN ARRAY
'''sum_of_weighted_y_values = 0
for i in y:
    sum_of_weighted_y_values = sum_of_weighted_y_values + i * w_i[numpy.where(y == i)]

y_weighted_mean = sum_of_weighted_y_values / numpy.sum([w_i])

sum_of_weighted_x_values = 0
for i in x:
    sum_of_weighted_x_values = sum_of_weighted_x_values + i * w_i[numpy.where(x == i)]

x_weighted_mean = sum_of_weighted_x_values / numpy.sum([w_i])

# Computes the sum of squared deviation in x and y values from the mean values
sum_of_x_errors_squared = 0
for i in y:
    y_error = i - y_weighted_mean
    y_error_squared = y_error ** 2
    sum_of_x_errors_squared = sum_of_x_errors_squared + y_error_squared * w_i[numpy.where(y == i)]

sum_of_x_errors_times_y_errors = 0
for i in x:
    y_at_x = y[numpy.where(x == i)]
    x_error = i - x_weighted_mean
    y_error = y_at_x - y_weighted_mean
    x_error_times_y_error = x_error * y_error
    sum_of_x_errors_times_y_errors = sum_of_x_errors_times_y_errors + x_error_times_y_error * w_i[numpy.where(x == i)]

# Performs weighted least-squares fit
density = sum_of_x_errors_times_y_errors / sum_of_x_errors_squared
print (density)

#ORIGINAL DATAPOINT POSITION WEIGHTS -- too much overlap, too hardcoded
for i in range(0, 400):
        if 0 <= i <= 19 or 380 <= i <= 399 or i % 20 == 0 or i % 20 == 19:
             datapoint_position_weight = numpy.append(datapoint_position_weight, 0.1)

        if 21 <= i <= 38 or 361 <= i <= 378 or i % 20 == 1 or i % 20 == 18:
            datapoint_position_weight = numpy.append(datapoint_position_weight, 0.2)

        if 42 <= i <= 57 or 342 <= i <= 357 or i % 20 == 2 or i % 20 == 17:
            datapoint_position_weight = numpy.append(datapoint_position_weight, 0.3)

        if 63 <= i <= 76 or 323 <= i <= 337 or i % 20 == 3 or i % 20 == 16:
            datapoint_position_weight = numpy.append(datapoint_position_weight, 0.4)

        if 84 <= i <= 95 or 304 <= i <= 315 or i % 20 == 4 or i % 20 == 15:
            datapoint_position_weight = numpy.append(datapoint_position_weight, 0.5)

        if 105 <= i <= 114 or 285 <= i <= 294 or i % 20 == 5 or i % 20 == 14:
            datapoint_position_weight = numpy.append(datapoint_position_weight, 0.6)

        if 126 <= i <= 134 or 266 <= i <= 273 or i % 20 == 6 or i % 20 == 13:
            datapoint_position_weight = numpy.append(datapoint_position_weight, 0.7)

        if 147 <= i <= 152 or 247 <= i <= 252 or i % 20 == 7 or i % 20 == 12:
            datapoint_position_weight = numpy.append(datapoint_position_weight, 0.8)

        if 168 <= i <= 171 or 228 <= i <= 231 or i % 20 == 8 or i % 20 == 11:
            datapoint_position_weight = numpy.append(datapoint_position_weight, 0.9)

        if i == 189 or i == 190 or i == 209 or i == 210:
            datapoint_position_weight = numpy.append(datapoint_position_weight, 1.0)
print (datapoint_position_weight)


'''
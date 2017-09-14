import numpy
import openpyxl
import math
from openpyxl.cell import column_index_from_string
from openpyxl.cell import get_column_letter

#Opens workbooks and takes basin name
topo_wb = openpyxl.load_workbook("Edited Topography.xlsx", read_only= True)
topo_sheet = topo_wb.get_sheet_by_name("Topography")
basin_name = input ("Input the name of the basin you wish to study and press the enter key")
center_excel_coordinates = input("Input the coordinates of the center of the basin you wish to study")
basin_radius = float(input("Input the radius of the basin in Excel coordinates (ten times the radius in degrees)"))
rim_radius = float(input("Input the radius of the rim in Excel coordinates (ten times the radius in degrees)"))
exterior_radius = float(input("Input the radius of the exterior in Excel coordinates (ten times the radius in degrees)"))
search_area_top_left = input("Input the top left Excel coordinates of the rectangular search area")
search_area_bottom_right = input("Input the bottom right Excel coordinates of the rectangular search area")

center_row = int(topo_sheet.cell(center_excel_coordinates).row)
center_column = int(topo_sheet.cell(center_excel_coordinates).column)
topo_data_input = tuple(topo_sheet[search_area_top_left: search_area_bottom_right])

basin_data = numpy.array([])
rim_data = numpy.array([])
exterior_data = numpy.array([])

for rowOfCellObjects in topo_data_input:
    for cellObj in rowOfCellObjects:
        column = cellObj.column
        column_from_center = column - center_column
        row = cellObj.row
        row_from_center = row - center_row
        topography_value = cellObj.value
        if math.sqrt((row_from_center ** 2) + (column_from_center ** 2)) <= basin_radius:
            basin_data = numpy.append(basin_data, topography_value)
        if basin_radius < math.sqrt((row_from_center ** 2) + (column_from_center ** 2)) <= rim_radius:
            rim_data = numpy.append(rim_data, topography_value)
        if rim_radius < math.sqrt((row_from_center ** 2) + (column_from_center ** 2)) <= exterior_radius:
            exterior_data = numpy.append(exterior_data, topography_value)

#print(basin_data)
#print(rim_data)
#print(exterior_data)

average_basin_topography = numpy.mean(basin_data)
average_rim_topography = numpy.mean(rim_data)
average_exterior_topography = numpy.mean(exterior_data)

print("The average topography in the " + basin_name + " basin is " + str(average_basin_topography) + " meters above the center of the Moon.")
print("The average topography in the " + basin_name + " rim is " + str(average_rim_topography) + " meters above the center of the Moon.")
print("The average topography in the " + basin_name + " exterior is " + str(average_exterior_topography) + " meters above the center of the Moon.")


#Creates coordinates for rectangular basin region
'''basin_top_left = input("Input the coordinates of the top-left corner of the basin area you wish to study and press the enter key")
basin_bottom_right = input("Input the coordinates of the bottom-right corner of basin area you wish to study and press the enter key")
rim_top_left = input("Input the coordinates of the top-left corner of the basin rim area you wish to study and press the enter key")
rim_bottom_right = input("Input the coordinates of the bottom-right corner of basin rim area you wish to study and press the enter key")
exterior_top_left = input("Input the coordinates of the top-left corner of the basin exterior area you wish to study and press the enter key")
exterior_bottom_right = input("Input the coordinates of the bottom-right corner of basin exterior area you wish to study and press the enter key")

basin_top_row = int(topo_sheet.cell(basin_top_left).row)
basin_left_column = int(topo_sheet.cell(basin_top_left).column)
basin_bottom_row = int(topo_sheet.cell(basin_bottom_right).row)
basin_right_column = int(topo_sheet.cell(basin_bottom_right).column)

basin_left_column = get_column_letter(basin_left_column)
basin_top_left = str(str(basin_left_column) + str(basin_top_row))
basin_right_column = get_column_letter(basin_right_column)
basin_bottom_right = str(str(basin_right_column) + str(basin_bottom_row))

#Inputs topography data in the basin and calculates the average value of topography in the basin
topo_basin_data_input = tuple(topo_sheet[basin_top_left : basin_bottom_right])

basin_topography = numpy.array([])
for rowOfCellObjects in topo_basin_data_input:
    for cellObj in rowOfCellObjects:
        topography_data = cellObj.value
        basin_topography = numpy.append(basin_topography, topography_data)

average_basin_topography = numpy.mean(basin_topography)
print("The average topography in the " + basin_name + " basin is " + str(average_basin_topography) + " meters above the center of the Moon.")
print(len(basin_topography))

#Creates coordinates for rectangular rim region
rim_top_row = int(topo_sheet.cell(rim_top_left).row)
rim_left_column = int(topo_sheet.cell(rim_top_left).column)
rim_bottom_row = int(topo_sheet.cell(rim_bottom_right).row)
rim_right_column = int(topo_sheet.cell(rim_bottom_right).column)

rim_left_column = get_column_letter(rim_left_column)
rim_top_left = str(str(rim_left_column) + str(rim_top_row))
rim_right_column = get_column_letter(rim_right_column)
rim_bottom_right = str(str(rim_right_column) + str(rim_bottom_row))
rim_top_side_bottom_right = str(str(rim_right_column) + str((basin_top_row - 1)))
rim_left_side_top_left = str(str(rim_left_column) + str(basin_top_row))
rim_left_side_bottom_right = str(str(get_column_letter(topo_sheet.cell(basin_top_left).column - 1)) + str(basin_bottom_row))
rim_bottom_side_top_left = str(str(rim_left_column) + str(basin_bottom_row + 1))
rim_right_side_top_left = str(str(get_column_letter(topo_sheet.cell(basin_bottom_right).column + 1)) + str(basin_top_row))
rim_right_side_bottom_right = str(str(rim_right_column) + str(basin_bottom_row))
rim_top_side = tuple(topo_sheet[rim_top_left : rim_top_side_bottom_right])
rim_left_side = tuple (topo_sheet[rim_left_side_top_left : rim_left_side_bottom_right])
rim_bottom_side = tuple(topo_sheet[rim_bottom_side_top_left : rim_bottom_right])
rim_right_side = tuple(topo_sheet[rim_right_side_top_left : rim_right_side_bottom_right])
topo_rim_data_input = rim_top_side + rim_left_side + rim_bottom_side + rim_right_side

#Inputs topography data in the rim and calculates the average value of topography in the rim
rim_topography = numpy.array([])
for rowOfCellObjects in topo_rim_data_input:
    for cellObj in rowOfCellObjects:
        topography_data = cellObj.value
        rim_topography = numpy.append(rim_topography, topography_data)

average_rim_topography = numpy.mean(rim_topography)
print("The average topography in the " + basin_name + " rim is " + str(average_rim_topography) + " meters above the center of the Moon.")
print(len(rim_topography))

#Creates coordinates for rectangular exterior region
exterior_top_row = int(topo_sheet.cell(exterior_top_left).row)
exterior_left_column = int(topo_sheet.cell(exterior_top_left).column)
exterior_bottom_row = int(topo_sheet.cell(exterior_bottom_right).row)
exterior_right_column = int(topo_sheet.cell(exterior_bottom_right).column)

exterior_left_column = get_column_letter(exterior_left_column)
exterior_top_left = str(str(exterior_left_column) + str(exterior_top_row))
exterior_right_column = get_column_letter(exterior_right_column)
exterior_bottom_right = str(str(exterior_right_column) + str(exterior_bottom_row))
exterior_top_side_bottom_right = str(str(exterior_right_column) + str((rim_top_row - 1)))
exterior_left_side_top_left = str(str(exterior_left_column) + str(rim_top_row))
exterior_left_side_bottom_right = str(str(get_column_letter(topo_sheet.cell(rim_top_left).column - 1)) + str(rim_bottom_row))
exterior_bottom_side_top_left = str(str(exterior_left_column) + str(rim_bottom_row + 1))
exterior_right_side_top_left = str(str(get_column_letter(topo_sheet.cell(rim_bottom_right).column + 1)) + str(rim_top_row))
exterior_right_side_bottom_right = str(str(exterior_right_column) + str(rim_bottom_row))
exterior_top_side = tuple(topo_sheet[exterior_top_left : exterior_top_side_bottom_right])
exterior_left_side = tuple (topo_sheet[exterior_left_side_top_left : exterior_left_side_bottom_right])
exterior_bottom_side = tuple(topo_sheet[exterior_bottom_side_top_left : exterior_bottom_right])
exterior_right_side = tuple(topo_sheet[exterior_right_side_top_left : exterior_right_side_bottom_right])
topo_exterior_data_input = exterior_top_side + exterior_left_side + exterior_bottom_side + exterior_right_side

#Inputs topography data in the exterior and calculates the average value of topography in the exterior
exterior_topography = numpy.array([])
for rowOfCellObjects in topo_exterior_data_input:
    for cellObj in rowOfCellObjects:
        topography_data = cellObj.value
        exterior_topography = numpy.append(exterior_topography, topography_data)

average_exterior_topography = numpy.mean(exterior_topography)
print("The average topography in the " + basin_name + " exterior is " + str(average_exterior_topography) + " meters above the center of the Moon.")
print(len(exterior_topography))'''


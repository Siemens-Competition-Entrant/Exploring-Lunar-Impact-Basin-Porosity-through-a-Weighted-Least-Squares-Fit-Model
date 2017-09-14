import numpy
import openpyxl
import math
from openpyxl.cell import column_index_from_string
from openpyxl.cell import get_column_letter
from openpyxl.cell import Cell

#Takes input of region and needed values
region_name = input ("Input the name of the region you wish to study and press the enter key")
type_of_data = input("What kind of files: Basin, Rim, or Exterior?")
'''grav_wb_name = input("Input the name of the gravity file you plan to use")
grav_sheet_name = input("Input the name of the gravity sheet you plan to use")
topo_wb_name = input("Input the name of the gravity-from-topography file you plan to use")
topo_sheet_name = input("Input the name of the gravity-from-topography sheet you plan to use")
grain_density_wb_name = input("Input the name of the grain density file you plan to use")
grain_density_sheet_name = input("Input the name of the grain density sheet you plan to use")'''
center_excel_coordinates = input("Input the coordinates of the center of the basin you wish to study")
basin_radius = float(input("Input the radius of the basin in Excel coordinates"))
rim_radius = float(input("Input the radius of the rim in Excel coordinates"))
exterior_radius = float(input("Input the radius of the exterior in Excel coordinates"))
search_area_top_left = input("Input the top left Excel coordinates of the rectangular search area")
search_area_bottom_right = input("Input the bottom right Excel coordinates of the rectangular search area")
#number_of_km_per_pixel = float(input("Input the number of kilometers per Excel pixel"))
number_of_km_per_pixel = 5

#Opens Excel datasets
'''grav_wb = openpyxl.load_workbook(grav_wb_name + ".xlsx", read_only=True)
topo_wb = openpyxl.load_workbook(topo_wb_name + ".xlsx", read_only= True)
grain_density_wb = openpyxl.load_workbook(grain_density_wb_name + ".xlsx", read_only= True)
grav_sheet = grav_wb.get_sheet_by_name(grav_sheet_name)
topo_sheet = topo_wb.get_sheet_by_name(topo_sheet_name)
grain_density_sheet = grain_density_wb.get_sheet_by_name(grain_density_sheet_name)'''

grav_wb = openpyxl.load_workbook(region_name + "_" + type_of_data + "Gravity.xlsx", read_only=True)
topo_wb = openpyxl.load_workbook(region_name + "_" + type_of_data + "GravityFromTopography.xlsx", read_only=True)
grain_density_wb = openpyxl.load_workbook(region_name  + "_GrainDensity.xlsx", read_only=True)
grav_sheet = grav_wb.get_sheet_by_name("Sheet1")
topo_sheet = topo_wb.get_sheet_by_name("Sheet1")
grain_density_sheet = grain_density_wb.get_sheet_by_name("Sheet1")

#Inputs data from datasets
center_row = int(topo_sheet.cell(center_excel_coordinates).row)
center_column = int(topo_sheet.cell(center_excel_coordinates).column)
grav_data_input = tuple(grav_sheet[search_area_top_left: search_area_bottom_right])
topo_data_input = tuple(topo_sheet[search_area_top_left: search_area_bottom_right])
grain_density_data_input = tuple(grain_density_sheet[search_area_top_left: search_area_bottom_right])

#Creates arrays to store data from datasets
basin_y_data = numpy.array([])
basin_x_data = numpy.array([])
basin_grain_density_data = numpy.array([])

rim_y_data = numpy.array([])
rim_x_data = numpy.array([])
rim_grain_density_data = numpy.array([])

exterior_y_data = numpy.array([])
exterior_x_data = numpy.array([])
exterior_grain_density_data = numpy.array([])

# Inputs gravity data of basin, rim, and exterior
for rowOfCellObjects in grav_data_input:
    for cellObj in rowOfCellObjects:
        column = cellObj.column
        column_from_center = column - center_column
        row = cellObj.row
        row_from_center = row - center_row
        gravity_value = cellObj.value
        if math.sqrt((row_from_center ** 2) + (column_from_center ** 2)) <= basin_radius:
            basin_y_data = numpy.append(basin_y_data, gravity_value)
        if basin_radius < math.sqrt((row_from_center ** 2) + (column_from_center ** 2)) <= rim_radius:
            rim_y_data = numpy.append(rim_y_data, gravity_value)
        if rim_radius < math.sqrt((row_from_center ** 2) + (column_from_center ** 2)) <= exterior_radius:
            exterior_y_data = numpy.append(exterior_y_data, gravity_value)

print(len(basin_y_data))
print(len(exterior_y_data))

# Inputs gravity-from-topography data of basin, rim, and exterior
for rowOfCellObjects in topo_data_input:
    for cellObj in rowOfCellObjects:
        column = cellObj.column
        column_from_center = column - center_column
        row = cellObj.row
        row_from_center = row - center_row
        topography_value = cellObj.value
        if math.sqrt((row_from_center ** 2) + (column_from_center ** 2)) <= basin_radius:
            basin_x_data = numpy.append(basin_x_data, topography_value)
        if basin_radius < math.sqrt((row_from_center ** 2) + (column_from_center ** 2)) <= rim_radius:
            rim_x_data = numpy.append(rim_x_data, topography_value)
        if rim_radius < math.sqrt((row_from_center ** 2) + (column_from_center ** 2)) <= exterior_radius:
            exterior_x_data = numpy.append(exterior_x_data, topography_value)

#Performs a least squares fit to find density in the basin
m_basin, b_basin = numpy.polyfit(basin_x_data, basin_y_data, 1)
m_basin = round(m_basin, 3)
b_basin = round(b_basin, 3)
basin_density = m_basin

#Standard error of basin density
#Finds error in x (gravity-from-topography) values and squares it
basin_x_mean = numpy.mean(basin_x_data)
basin_sum_of_x_errors_squared = 0
for i in basin_x_data:
    x_error = i - basin_x_mean
    x_error_squared = x_error ** 2
    basin_sum_of_x_errors_squared = basin_sum_of_x_errors_squared + x_error_squared

# Computes the sum of squared differences in the actual y-value from the predicted y-value
basin_differences_from_regression_model = numpy.array([])
basin_squared_differences_from_predicted = numpy.array([])
basin_best_fit_error = 0
basin_difference_from_predicted = basin_y_data - (m_basin * basin_x_data + b_basin)
basin_differences_from_regression_model = numpy.append(basin_differences_from_regression_model, basin_difference_from_predicted)
for i in basin_differences_from_regression_model:
    differences_squared = i ** 2
    basin_squared_differences_from_predicted = numpy.append(basin_squared_differences_from_predicted, differences_squared)
    basin_best_fit_error = basin_best_fit_error + differences_squared

#Computes patch area for the basin and calculates standard error
basin_radius_km = basin_radius * number_of_km_per_pixel
basin_patch_area = math.pi * (basin_radius_km ** 2) #Change this to reflect radius in km
basin_standard_error = math.sqrt(basin_best_fit_error / basin_sum_of_x_errors_squared) / ((basin_patch_area * 241001 / (3.8 * 10 ** 7))) #ASK DR. JAMES
basin_standard_error = round(basin_standard_error, 4)
basin_confidence_interval = basin_standard_error * 2

#Performs a least squares fit to find density in the rim
m_rim, b_rim = numpy.polyfit(rim_x_data, rim_y_data, 1)
m_rim = round(m_rim, 3)
b_rim = round(b_rim, 3)
rim_density = m_rim

#Standard error of rim density
#Finds error in x (gravity-from-topography) values and squares it
rim_x_mean = numpy.mean(rim_x_data)
rim_sum_of_x_errors_squared = 0
for i in rim_x_data:
    x_error = i - rim_x_mean
    x_error_squared = x_error ** 2
    rim_sum_of_x_errors_squared = rim_sum_of_x_errors_squared + x_error_squared

# Computes the sum of squared differences in the actual y-value from the predicted y-value
rim_differences_from_regression_model = numpy.array([])
rim_squared_differences_from_predicted = numpy.array([])
rim_best_fit_error = 0
rim_difference_from_predicted = rim_y_data - (m_rim * rim_x_data + b_rim)
rim_differences_from_regression_model = numpy.append(rim_differences_from_regression_model, rim_difference_from_predicted)
for i in rim_differences_from_regression_model:
    differences_squared = i ** 2
    rim_squared_differences_from_predicted = numpy.append(rim_squared_differences_from_predicted, differences_squared)
    rim_best_fit_error = rim_best_fit_error + differences_squared

#Computes patch area for the rim and calculates standard error
rim_radius_km = rim_radius * number_of_km_per_pixel
rim_patch_area = math.pi * (rim_radius_km ** 2) - basin_patch_area #Change this to reflect radius in km
rim_standard_error = math.sqrt(rim_best_fit_error / rim_sum_of_x_errors_squared) / ((rim_patch_area * 241001 / (3.8 * 10 ** 7))) #ASK DR. JAMES
rim_standard_error = round(rim_standard_error, 4)
rim_confidence_interval = rim_standard_error * 2

#Performs a least squares fit to find density in the exterior
m_exterior, b_exterior = numpy.polyfit(exterior_x_data, exterior_y_data, 1)
m_exterior = round(m_exterior, 3)
b_exterior = round(b_exterior, 3)
exterior_density = m_exterior

#Standard error of exterior density
#Finds error in x (gravity-from-topography) values and squares it
exterior_x_mean = numpy.mean(exterior_x_data)
exterior_sum_of_x_errors_squared = 0
for i in exterior_x_data:
    x_error = i - exterior_x_mean
    x_error_squared = x_error ** 2
    exterior_sum_of_x_errors_squared = exterior_sum_of_x_errors_squared + x_error_squared

# Computes the sum of squared differences in the actual y-value from the predicted y-value
exterior_differences_from_regression_model = numpy.array([])
exterior_squared_differences_from_predicted = numpy.array([])
exterior_best_fit_error = 0
exterior_difference_from_predicted = exterior_y_data - (m_exterior * exterior_x_data + b_exterior)
exterior_differences_from_regression_model = numpy.append(exterior_differences_from_regression_model,
                                                       exterior_difference_from_predicted)
for i in exterior_differences_from_regression_model:
    differences_squared = i ** 2
    exterior_squared_differences_from_predicted = numpy.append(exterior_squared_differences_from_predicted, differences_squared)
    exterior_best_fit_error = exterior_best_fit_error + differences_squared

#Computes patch area for the exterior and calculates standard error
exterior_radius_km = exterior_radius * number_of_km_per_pixel
exterior_patch_area = math.pi * (exterior_radius_km ** 2) - rim_patch_area #Change this to reflect radius in km
exterior_standard_error = math.sqrt(exterior_best_fit_error / exterior_sum_of_x_errors_squared) / ((exterior_patch_area * 234726 / (3.8 * 10 ** 7)))
exterior_standard_error = round(exterior_standard_error, 4)
exterior_confidence_interval = exterior_standard_error * 2


# Computes the area of the selected region on the Moon in old standard error calculation
'''basin_min_row = grav_sheet.cell(top_left_coordinates).row
basin_min_column = grav_sheet.cell(top_left_coordinates).column
basin_max_row = grav_sheet.cell(bottom_right_coordinates).row
basin_max_column = grav_sheet.cell(bottom_right_coordinates).column

rows = max_row - min_row + 1
columns = max_column - min_column + 1

latitude_height = rows / 10
longitude_width = columns / 10

pi = math.pi
patch_area = latitude_height * longitude_width * ((1738 * pi / 180) ** 2) * math.cos(median_latitude * pi / 180)'''

#Inputs grain density values of basin, rim, and exterior
for rowOfCellObjects in grain_density_data_input:
    for cellObj in rowOfCellObjects:
        column = cellObj.column
        column_from_center = column - center_column
        row = cellObj.row
        row_from_center = row - center_row
        grain_density_value = cellObj.value / 1000
        if math.sqrt((row_from_center ** 2) + (column_from_center ** 2)) <= basin_radius:
            basin_grain_density_data = numpy.append(basin_grain_density_data, grain_density_value)
        if basin_radius < math.sqrt((row_from_center ** 2) + (column_from_center ** 2)) <= rim_radius:
            rim_grain_density_data = numpy.append(rim_grain_density_data, grain_density_value)
        if rim_radius < math.sqrt((row_from_center ** 2) + (column_from_center ** 2)) <= exterior_radius:
            exterior_grain_density_data = numpy.append(exterior_grain_density_data, grain_density_value)

#Calculates mean grain density
basin_grain_density_mean = numpy.mean(basin_grain_density_data)
rim_grain_density_mean = numpy.mean(rim_grain_density_data)
exterior_grain_density_mean = numpy.mean(exterior_grain_density_data)

# Calculates porosity and confidence interval based on density and grain density
#Calculates porosity as a decimal
basin_porosity = 1 - basin_density / basin_grain_density_mean
rim_porosity = 1 - rim_density / rim_grain_density_mean
exterior_porosity = 1 - exterior_density / exterior_grain_density_mean

#Converts decimal porosity to percent
basin_percent_porosity = basin_porosity * 100
basin_percent_porosity = round(basin_percent_porosity, 3)
rim_percent_porosity = rim_porosity * 100
rim_percent_porosity = round(rim_percent_porosity, 3)
exterior_percent_porosity = exterior_porosity * 100
exterior_percent_porosity = round(exterior_percent_porosity, 3)

#Porosity standard error and confidence interval.  Also multiplies porosity error by 100 to account for conversion of porosity into a percent
basin_porosity_error = basin_standard_error / basin_grain_density_mean
basin_percent_porosity_error = basin_porosity_error * 100
basin_percent_porosity_error = round(basin_percent_porosity_error, 3)
rim_porosity_error = rim_standard_error / rim_grain_density_mean
rim_percent_porosity_error = rim_porosity_error * 100
rim_percent_porosity_error = round(rim_percent_porosity_error, 3)
exterior_porosity_error = exterior_standard_error / exterior_grain_density_mean
exterior_percent_porosity_error = exterior_porosity_error * 100
exterior_percent_porosity_error = round(exterior_percent_porosity_error, 3)
basin_porosity_confidence_interval = basin_percent_porosity_error * 2
rim_porosity_confidence_interval = rim_percent_porosity_error * 2
exterior_porosity_confidence_interval = exterior_percent_porosity_error * 2

basin_porosity_with_confidence_interval = str(basin_percent_porosity) + " ± " + str(basin_porosity_confidence_interval)
rim_porosity_with_confidence_interval = str(rim_percent_porosity) + " ± " + str(rim_porosity_confidence_interval)
exterior_porosity_with_confidence_interval = str(exterior_percent_porosity) + " ± " + str(exterior_porosity_confidence_interval)

#Prints average basin, rim, and exterior porosity.
print (region_name + " Basin porosity = " + basin_porosity_with_confidence_interval + " (using " + type_of_data + " data)")
print (region_name + " Rim porosity = " + rim_porosity_with_confidence_interval + " (using " + type_of_data + " data)")
print (region_name + " Exterior porosity = " + exterior_porosity_with_confidence_interval + " (using " + type_of_data + " data)")


